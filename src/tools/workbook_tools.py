"""
Workbook management tools for Excel MCP server.
"""

import logging
from ..core.tag_system import ToolTags
from ..core.file_manager import get_safe_filename
from ..utils.workbook import get_workbook_info
from ..utils.exceptions import WorkbookError, ValidationError

logger = logging.getLogger("excel-mcp")


def register_workbook_tools(mcp_server):
    """Register all workbook-related tools with the MCP server."""
    
    @mcp_server.tool(
        tags=ToolTags(
            functional_domains=["excel"],
            permissions=["write"],
            resource_types=["workbook", "file"],
            scopes=["user_scoped"]
        )
    )
    def create_workbook(user_id: str, filename: str) -> str:
        """
        Create a new Excel workbook.
        
        Args:
            user_id: User ID for file organization
            filename: Name of the workbook file to create
            
        Returns:
            Success message with the created filename
        """
        try:
            # Ensure we only work with safe filenames
            safe_filename = get_safe_filename(filename)
            file_path = mcp_server.file_manager.get_file_path(safe_filename, user_id)
            
            with mcp_server.file_manager.lock_file(file_path):
                from openpyxl import Workbook
                
                # Create new workbook
                wb = Workbook()
                wb.save(str(file_path))
                
                logger.info(f"Created workbook: {safe_filename} for user {user_id}")
                return f"Workbook '{safe_filename}' created successfully"
                
        except Exception as e:
            logger.error(f"Error creating workbook: {e}")
            raise WorkbookError(f"Failed to create workbook: {str(e)}")
    
    @mcp_server.tool(
        tags=ToolTags(
            functional_domains=["excel"],
            permissions=["write"],
            resource_types=["sheet", "workbook"],
            scopes=["user_scoped"]
        )
    )
    def create_worksheet(user_id: str, filename: str, sheet_name: str) -> str:
        """
        Create a new worksheet in an existing workbook.
        
        Args:
            user_id: User ID for file organization
            filename: Name of the workbook file
            sheet_name: Name for the new worksheet
            
        Returns:
            Success message with the created sheet name
        """
        try:
            safe_filename = get_safe_filename(filename)
            file_path = mcp_server.file_manager.get_file_path(safe_filename, user_id)
            
            with mcp_server.file_manager.lock_file(file_path):
                from openpyxl import load_workbook
                
                wb = load_workbook(str(file_path))
                
                # Check if sheet already exists
                if sheet_name in wb.sheetnames:
                    return f"Error: Sheet '{sheet_name}' already exists"
                
                # Create new worksheet
                wb.create_sheet(title=sheet_name)
                wb.save(str(file_path))
                
                logger.info(f"Created worksheet '{sheet_name}' in {safe_filename} for user {user_id}")
                return f"Worksheet '{sheet_name}' created successfully in '{safe_filename}'"
                
        except Exception as e:
            logger.error(f"Error creating worksheet: {e}")
            raise WorkbookError(f"Failed to create worksheet: {str(e)}")
    
    @mcp_server.tool(
        tags=ToolTags(
            functional_domains=["excel"],
            permissions=["read"],
            resource_types=["workbook"],
            scopes=["user_scoped"]
        )
    )
    def get_workbook_metadata(
        user_id: str, 
        filename: str, 
        include_ranges: bool = False
    ) -> str:
        """
        Get metadata about workbook including sheets and ranges.
        
        Args:
            user_id: User ID for file organization
            filename: Name of the workbook file
            include_ranges: Whether to include range information
            
        Returns:
            JSON string with workbook metadata
        """
        try:
            safe_filename = get_safe_filename(filename)
            file_path = mcp_server.file_manager.get_file_path(safe_filename, user_id)
            
            # Use shared lock for read operation
            with mcp_server.file_manager.lock_file(file_path):
                result = get_workbook_info(str(file_path), include_ranges)
                
                # Return just the filename in metadata, not full path
                if isinstance(result, dict) and 'file_path' in result:
                    result['filename'] = safe_filename
                    del result['file_path']  # Remove full path to avoid confusion
                
                import json
                return json.dumps(result, indent=2, default=str)
                
        except Exception as e:
            logger.error(f"Error getting workbook metadata: {e}")
            raise WorkbookError(f"Failed to get workbook metadata: {str(e)}")