import logging
import json
from typing import Optional
from openpyxl import load_workbook
from ..core.file_manager import get_safe_file_name
from ..utils.data import read_excel_range_with_metadata
from ..utils.validation import validate_formula_in_cell_operation as validate_formula_impl
from ..utils.validation import validate_range_in_sheet_operation as validate_range_impl
from ..utils.calculations import CalculationError
from ..utils.cell_validation import get_all_validation_ranges
from ..utils.sheet import get_merged_ranges
from ..utils.workbook import get_workbook_info
from ..utils.exceptions import ValidationError, SheetError, WorkbookError

logger = logging.getLogger("excel-mcp")

def register_excel_read_tools(mcp_server):
    """Register all Excel read tools with the MCP server."""

    # Data related tools
    @mcp_server.tool(tags={"excel", "read"})
    def read_data_from_excel(
        user_id: str,
        file_name: str,
        sheet_name: str,
        start_cell: str,
        end_cell: str,
        preview_only: bool = False
    ) -> str:
        """
        Read data from Excel worksheet with cell metadata including validation rules.
        
        Args:
            user_id: User ID for file organization
            file_name: Name of the Excel file
            sheet_name: Name of worksheet
            start_cell: Starting cell
            end_cell: Ending cell
            preview_only: Whether to return preview only
        
        Returns:  
            JSON string containing structured cell data with validation metadata.
            Each cell includes: address, value, row, column, and validation info (if any).
        """
        safe_file_name = get_safe_file_name(file_name)
        file_path = mcp_server.file_manager.get_file_path(safe_file_name, user_id)
        try:
            with mcp_server.file_manager.lock_file(file_path):
                result = read_excel_range_with_metadata(
                    str(file_path), 
                    sheet_name, 
                    start_cell, 
                    end_cell
                )
                if not result or not result.get("cells"):
                    return "No data found in specified range"
                return json.dumps(result, indent=2, default=str)
        except Exception as e:
            logger.error(f"Error reading data: {e}")
            raise
        
    @mcp_server.tool(tags={"excel", "read"})
    def validate_formula_syntax(
        user_id: str,
        file_name: str,
        sheet_name: str,
        cell: str,
        formula: str,
    ) -> str:
        """
        Validate Excel formula syntax without applying it.
        
        Args:
            user_id: User ID for file organization
            file_name: Name of the Excel file
            sheet_name: Name of worksheet
            cell: Target cell address (e.g., 'A1')
            formula: Excel formula to validate
            
        Returns:
            Validation result message indicating if the formula is valid or not
        """
        safe_file_name = get_safe_file_name(file_name)
        file_path = mcp_server.file_manager.get_file_path(safe_file_name, user_id)
        try:
            with mcp_server.file_manager.lock_file(file_path):
                result = validate_formula_impl(str(file_path), sheet_name, cell, formula)
                safe_result = result["message"].replace(str(file_path), safe_file_name)
                return safe_result
        except (ValidationError, CalculationError) as e:
            safe_error = str(e).replace(str(file_path), safe_file_name)
            return f"Error: {safe_error}"
        except Exception as e:
            logger.error(f"Error validating formula: {e}")
            raise

    # Formatting tools
    @mcp_server.tool(tags={"excel", "read"})
    def validate_excel_range(
        user_id: str,
        file_name: str,
        sheet_name: str,
        start_cell: str,
        end_cell: str
    ) -> str:
        """
        Validate if a range exists and is properly formatted.
        
        Args:
            user_id (str): User ID for file organization
            file_name (str): Name of the Excel file
            sheet_name (str): Name of worksheet
            start_cell (str): Starting cell of range to validate (e.g., "A1")
            end_cell (str): Ending cell of range to validate (e.g., "C3").
                If not provided, only the start_cell is validated. Defaults to None.
                
        Returns:
            str: Success message with file_name
        """
        safe_file_name = get_safe_file_name(file_name)
        file_path = mcp_server.file_manager.get_file_path(safe_file_name, user_id)
        try:
            with mcp_server.file_manager.lock_file(file_path):
                result = validate_range_impl(str(file_path), sheet_name, start_cell, end_cell)
                safe_result = result["message"].replace(str(file_path), f"'{safe_file_name}'")
                return safe_result
        except (ValidationError) as e:
            safe_error = str(e).replace(str(file_path), f"'{safe_file_name}'")
            return f"Error: {safe_error}"
        except Exception as e:
            logger.error(f"Error validating range: {e}")
            raise
        
    @mcp_server.tool(tags={"excel", "read"})
    def get_data_validation_info(
        user_id: str,
        file_name: str,
        sheet_name: str
    ) -> str:
        """
        Get all data validation rules in a worksheet.
        
        This tool helps identify which cell ranges have validation rules 
        and what types of validation are applied.
        
        Args:
            user_id (str): User ID for file organization
            file_name (str): Name of the Excel file
            sheet_name (str): Name of worksheet
            
        Returns:
            str: JSON string containing all validation rules in the worksheet
        """
        safe_file_name = get_safe_file_name(file_name)
        file_path = mcp_server.file_manager.get_file_path(safe_file_name, user_id)
        try:
            with mcp_server.file_manager.lock_file(file_path):
                wb = load_workbook(str(file_path), read_only=False)
                if sheet_name not in wb.sheetnames:
                    return f"Error: Sheet '{sheet_name}' not found"
                ws = wb[sheet_name]
                validations = get_all_validation_ranges(ws)
                wb.close()
                if not validations:
                    return "No data validation rules found in this worksheet"
                return json.dumps({
                    "file_name": safe_file_name,
                    "sheet_name": sheet_name,
                    "validation_rules": validations
                }, indent=2, default=str)
                
        except Exception as e:
            logger.error(f"Error getting validation info: {e}")
            raise
        
    # Sheet related tools
    @mcp_server.tool(tags={"excel", "read"})
    def get_merged_cells(user_id: str, file_name: str, sheet_name: str) -> str:
        """
        Get all merged cell ranges in the worksheet.
        
        Args:
            user_id: User ID for file organization
            file_name: Name of the Excel file
            sheet_name: Name of worksheet
            
        Returns:
            String with merged ranges information
        """
        safe_file_name = get_safe_file_name(file_name)
        file_path = mcp_server.file_manager.get_file_path(safe_file_name, user_id)
        try:
            with mcp_server.file_manager.lock_file(file_path):
                result = str(get_merged_ranges(str(file_path), sheet_name))
                safe_result = result.replace(str(file_path), f"'{safe_file_name}'")
                return safe_result
        except (ValidationError, SheetError) as e:
            safe_error = str(e).replace(str(file_path), f"'{safe_file_name}'")
            return f"Error: {safe_error}"
        except Exception as e:
            logger.error(f"Error getting merged cells: {e}")
            raise

    # Workbook related tools
    @mcp_server.tool(tags={"excel", "read"})
    def get_workbook_metadata(
        user_id: str, 
        file_name: str, 
        include_ranges: bool = False
    ) -> str:
        """
        Get metadata about workbook including sheets and ranges.
        
        Args:
            user_id: User ID for file organization
            file_name: Name of the workbook file
            include_ranges: Whether to include range information, default to false
            
        Returns:
            JSON string with workbook metadata
        """
        safe_file_name = get_safe_file_name(file_name)
        file_path = mcp_server.file_manager.get_file_path(safe_file_name, user_id)
        try:
            with mcp_server.file_manager.lock_file(file_path):
                result = get_workbook_info(str(file_path), include_ranges=include_ranges)
                # Normalize file_name to avoid leaking any path and align with other tools' outputs
                if isinstance(result, dict):
                    result["file_name"] = safe_file_name
                return json.dumps(result, indent=2, default=str)
        except WorkbookError as e:
            safe_error = str(e).replace(str(file_path), f"'{safe_file_name}'")
            return f"Error: {safe_error}"
        except Exception as e:
            logger.error(f"Error getting workbook metadata: {e}")
            raise