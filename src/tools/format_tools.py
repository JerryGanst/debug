"""
Formatting and validation tools for Excel MCP server.
"""

import logging
import json
from typing import Optional, Dict, Any
from ..core.tag_system import ToolTags
from ..core.file_manager import get_safe_filename
from ..utils.formatting import format_range as format_range_util
from ..utils.validation import validate_range_in_sheet_operation as validate_range_impl
from ..utils.cell_validation import get_all_validation_ranges
from ..utils.exceptions import FormattingError, ValidationError
from ..utils.sheet import copy_range_operation
from ..utils.sheet import delete_range_operation

logger = logging.getLogger("excel-mcp")


def register_format_tools(mcp_server):
    """Register all formatting and validation tools with the MCP server."""
    
    @mcp_server.tool(
        tags=ToolTags(
            functional_domains=["excel"],
            permissions=["write"],
            resource_types=["format", "cell"],
            scopes=["user_scoped"]
        )
    )
    def format_range(
        user_id: str,
        filename: str,
        sheet_name: str,
        start_cell: str,
        end_cell: Optional[str] = None,
        bold: bool = False,
        italic: bool = False,
        underline: bool = False,
        font_size: Optional[int] = None,
        font_color: Optional[str] = None,
        bg_color: Optional[str] = None,
        border_style: Optional[str] = None,
        border_color: Optional[str] = None,
        number_format: Optional[str] = None,
        alignment: Optional[str] = None,
        wrap_text: bool = False,
        merge_cells: bool = False,
        protection: Optional[Dict[str, Any]] = None,
        conditional_format: Optional[Dict[str, Any]] = None
    ) -> str:
        """
        Format a range of cells with various styling options.
        
        Args:
            user_id: User ID for file organization
            filename: Name of the Excel file
            sheet_name: Name of worksheet
            start_cell: Start cell of the range (e.g., 'A1')
            end_cell: End cell of the range (e.g., 'C10')
            bold: Whether text should be bold
            italic: Whether text should be italic
            underline: Whether text should be underlined
            font_size: Font size (e.g., 12)
            font_color: Font color (hex or name)
            bg_color: Background color (hex or name)
            border_style: Border style (e.g., 'thin', 'medium', 'thick')
            border_color: Border color (hex or name)
            number_format: Number format string
            alignment: Text alignment
            wrap_text: Whether text should wrap
            merge_cells: Whether to merge cells
            protection: Cell protection settings
            conditional_format: Conditional formatting rules
            
        Returns:
            Success message with filename
        """
        try:
            safe_filename = get_safe_filename(filename)
            file_path = mcp_server.file_manager.get_file_path(safe_filename, user_id)
            
            with mcp_server.file_manager.lock_file(file_path):
                
                format_range_util(
                    filepath=str(file_path),
                    sheet_name=sheet_name,
                    start_cell=start_cell,
                    end_cell=end_cell,  # This can be None
                    bold=bold,
                    italic=italic,
                    underline=underline,
                    font_size=font_size,  # This can be None
                    font_color=font_color,  # This can be None
                    bg_color=bg_color,  # This can be None
                    border_style=border_style,  # This can be None
                    border_color=border_color,  # This can be None
                    number_format=number_format,  # This can be None
                    alignment=alignment,  # This can be None
                    wrap_text=wrap_text,
                    merge_cells=merge_cells,
                    protection=protection,  # This can be None
                    conditional_format=conditional_format  # This can be None
                )
                
                return f"Range '{start_cell}':'{end_cell}' formatted successfully in sheet '{sheet_name}' of '{safe_filename}'"
                
        except (ValidationError, FormattingError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error formatting range: {e}")
            raise FormattingError(f"Failed to format range: {str(e)}")
        
    @mcp_server.tool(
        tags=ToolTags(
            functional_domains=["excel"],
            permissions=["write"],
            resource_types=["validation"],
            scopes=["user_scoped"]
        )
    )
    def copy_range(
        user_id: str,
        filename: str,
        sheet_name: str,
        source_start: str,
        source_end: str,
        target_start: str,
        target_sheet: Optional[str] = None
    ) -> str:
        """Copy a range of cells to another location."""
        try:
            safe_filename = get_safe_filename(filename)
            file_path = mcp_server.file_manager.get_file_path(safe_filename, user_id)
            result = copy_range_operation(
                full_path = str(file_path),
                sheet_name = sheet_name,
                source_start = source_start,
                source_end = source_end,
                target_start = target_start,
                target_sheet = target_sheet or sheet_name  # Use source sheet if target_sheet is None
            )
            return result["message"]
        except (ValidationError, SheetError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error copying range: {e}")
            raise
        
    @mcp_server.tool(
        tags=ToolTags(
            functional_domains=["excel"],
            permissions=["write"],
            resource_types=["validation"],
            scopes=["user_scoped"]
        )
    )
    def delete_range(
        user_id: str,
        filename: str,
        sheet_name: str,
        start_cell: str,
        end_cell: str,
        shift_direction: str = "up"
    ) -> str:
        """Delete a range of cells and shift remaining cells."""
        try:
            safe_filename = get_safe_filename(filename)
            file_path = mcp_server.file_manager.get_file_path(safe_filename, user_id)
            result = delete_range_operation(
                filepath=str(file_path),
                sheet_name=sheet_name,
                start_cell=start_cell,
                end_cell=end_cell,
                shift_direction=shift_direction
            )
            return result["message"]
        except (ValidationError, SheetError) as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error deleting range: {e}")
            raise
        
    @mcp_server.tool(
            tags=ToolTags(
                functional_domains=["excel"],
                permissions=["read"],
                resource_types=["validation"],
                scopes=["user_scoped"]
            )
        )
    def validate_excel_range(
        user_id: str,
        filename: str,
        sheet_name: str,
        start_cell: str,
        end_cell: Optional[str] = None
    ) -> str:
        """Validate if a range exists and is properly formatted."""
        try:
            safe_filename = get_safe_filename(filename)
            file_path = mcp_server.file_manager.get_file_path(safe_filename, user_id)
            range_str = start_cell if not end_cell else f"{start_cell}:{end_cell}"
            result = validate_range_impl(str(file_path), sheet_name, range_str)
            return result["message"]
        except ValidationError as e:
            return f"Error: {str(e)}"
        except Exception as e:
            logger.error(f"Error validating range: {e}")
            raise ValidationError(f"Failed to validate range: {str(e)}")

    @mcp_server.tool(
        tags=ToolTags(
            functional_domains=["excel"],
            permissions=["read"],
            resource_types=["validation"],
            scopes=["user_scoped"]
        )
    )
    def get_data_validation_info(
        user_id: str,
        filename: str,
        sheet_name: str
    ) -> str:
        """
        Get all data validation rules in a worksheet.
        
        This tool helps identify which cell ranges have validation rules 
        and what types of validation are applied.
        
        Args:
            user_id: User ID for file organization
            filename: Name of the Excel file
            sheet_name: Name of worksheet
            
        Returns:
            JSON string containing all validation rules in the worksheet
        """
        try:
            safe_filename = get_safe_filename(filename)
            file_path = mcp_server.file_manager.get_file_path(safe_filename, user_id)
            
            with mcp_server.file_manager.lock_file(file_path):
                from openpyxl import load_workbook
                
                wb = load_workbook(str(file_path), read_only=False)
                if sheet_name not in wb.sheetnames:
                    return f"Error: Sheet '{sheet_name}' not found"
                    
                ws = wb[sheet_name]
                validations = get_all_validation_ranges(ws)
                wb.close()
                
                if not validations:
                    return "No data validation rules found in this worksheet"
                    
                return json.dumps({
                    "filename": safe_filename,
                    "sheet_name": sheet_name,
                    "validation_rules": validations
                }, indent=2, default=str)
                
        except Exception as e:
            logger.error(f"Error getting validation info: {e}")
            raise ValidationError(f"Failed to get validation info: {str(e)}")