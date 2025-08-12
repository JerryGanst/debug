import uuid
import logging

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from .exceptions import DataError

logger = logging.getLogger(__name__)

def create_excel_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    table_name: str | None = None,
    table_style: str = "TableStyleMedium9"
) -> dict:
    """Creates a native Excel table for the given data range.
    
    Args:
        filepath: Path to the Excel file.
        sheet_name: Name of the worksheet.
        data_range: The cell range for the table (e.g., "A1:D5").
        table_name: A unique name for the table. If not provided, a unique name is generated.
        table_style: The visual style to apply to the table.
        
    Returns:
        A dictionary with a success message and table details.
    """
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise DataError(f"Sheet '{sheet_name}' not found.")
            
        ws = wb[sheet_name]

        # If no table name is provided, generate a unique one
        if not table_name:
            table_name = f"Table_{uuid.uuid4().hex[:8]}"

        # Collect existing table names across the entire workbook to ensure global uniqueness
        existing_table_names: set[str] = set()

        for sheet in wb.worksheets:
            # For openpyxl ≥3.1, Worksheet.tables is a dict {name: Table}
            if hasattr(sheet, "tables"):
                tables_attr = sheet.tables
                if isinstance(tables_attr, dict):
                    existing_table_names.update(tables_attr.keys())
                else:  # Older versions may expose a list/tuple of Table objects
                    existing_table_names.update(tbl.displayName for tbl in tables_attr)

            # Fallback: some older versions keep tables in the protected _tables list
            if hasattr(sheet, "_tables"):
                existing_table_names.update(tbl.displayName for tbl in sheet._tables)

        if table_name in existing_table_names:
            raise DataError(f"Table name '{table_name}' already exists in the workbook.")

        # Create the table
        table = Table(displayName=table_name, ref=data_range)
        
        # Apply style
        style = TableStyleInfo(
            name=table_style, 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        ws.add_table(table)
        
        wb.save(filepath)
        
        return {
            "message": f"Successfully created table '{table_name}' in sheet '{sheet_name}'.",
            "table_name": table_name,
            "range": data_range
        }

    except Exception as e:
        logger.error(f"Failed to create table: {e}")
        raise DataError(str(e)) 