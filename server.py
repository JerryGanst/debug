import logging
import os
from typing import Any, List, Dict, Optional
from fastmcp import FastMCP

# Import exceptions
from utils.exceptions import (
    ValidationError,
    WorkbookError,
    SheetError,
    DataError,
    FormattingError,
    CalculationError,
    PivotError,
    ChartError
)

# Import from util package with consistent _impl suffixes
from utils.validation import (
    validate_formula_in_cell_operation as validate_formula_impl,
    validate_range_in_sheet_operation as validate_range_impl
)
from utils.load_configs import load_configs
from utils.chart import create_chart_in_sheet as create_chart_impl
from utils.workbook import get_workbook_info
from utils.data import write_data
from utils.pivot import create_pivot_table as create_pivot_table_impl
from utils.tables import create_excel_table as create_table_impl
from utils.sheet import (
    copy_sheet,
    delete_sheet,
    rename_sheet,
    merge_range,
    unmerge_range,
    get_merged_ranges,
)

# 
from minio import Minio
from minio.error import S3Error
from pathlib import Path

# Get current directory for log file path.
LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel-mcp.log")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE)
    ],
)
logger = logging.getLogger("excel-mcp")
# Initialize FastMCP server (FastMCP 2.x)
mcp = FastMCP("excel-mcp")

def get_excel_path(filename: str, user_id: str = None) -> str:
    """Get full path to Excel file.
    
    Args:
        filename: Name of Excel file
        user_id: User ID for file organization (optional for backward compatibility)
        
    Returns:
        Full path to Excel file
    """
    # If filename is already an absolute path, return it
    if os.path.isabs(filename):
        return filename

    # Load configuration
    configs = load_configs()
    excel_files_path = configs.EXCEL_FILES_PATH
    
    # Resolve path based on EXCEL_FILES_PATH from configs
    if user_id:
        # Create user-specific directory path
        user_dir = os.path.join(excel_files_path, f"{user_id}")
        # Ensure the user directory exists
        os.makedirs(user_dir, exist_ok=True)
        return os.path.join(user_dir, filename)
    else:
        # Backward compatibility: use root EXCEL_FILES_PATH
        return os.path.join(excel_files_path, filename)

@mcp.tool()
def apply_formula(
    user_id: str,
    file_name: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """
    Apply Excel formula to cell.
    Excel formula will write to cell with verification.
    """
    try:
        full_path = get_excel_path(file_name, user_id)
        # First validate the formula
        validation = validate_formula_impl(full_path, sheet_name, cell, formula)
        if isinstance(validation, dict) and "error" in validation:
            return f"Error: {validation['error']}"
            
        # If valid, apply the formula
        from utils.calculations import apply_formula as apply_formula_impl
        result = apply_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error applying formula: {e}")
        raise

@mcp.tool()
def validate_formula_syntax(
    user_id: str,
    file_name: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """Validate Excel formula syntax without applying it."""
    try:
        full_path = get_excel_path(file_name, user_id)
        result = validate_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating formula: {e}")
        raise

@mcp.tool()
def format_range(
    user_id: str,
    file_name: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None,
    conditional_format: Optional[Dict[str, Any]] = None
) -> str:
    """Apply formatting to a range of cells."""
    try:
        full_path = get_excel_path(file_name, user_id)
        from utils.formatting import format_range as format_range_func
        
        # Convert None values to appropriate defaults for the underlying function
        format_range_func(
            filepath=full_path,
            sheet_name=sheet_name,
            start_cell=start_cell,
            end_cell=end_cell,  # This can be None
            bold=bold,
            italic=italic,
            underline=underline,
            font_size=font_size,  # This can be None
            font_color=font_color,  # This can be None
            bg_color=bg_color,  # This can be None
            border_style=border_style,  # This can be None
            border_color=border_color,  # This can be None
            number_format=number_format,  # This can be None
            alignment=alignment,  # This can be None
            wrap_text=wrap_text,
            merge_cells=merge_cells,
            protection=protection,  # This can be None
            conditional_format=conditional_format  # This can be None
        )
        return "Range formatted successfully"
    except (ValidationError, FormattingError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error formatting range: {e}")
        raise

@mcp.tool()
def read_data_from_excel(
    user_id: str,
    file_name: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    preview_only: bool = False
) -> str:
    """
    Read data from Excel worksheet with cell metadata including validation rules.
    
    Args:
        user_id: User ID for file organization
        file_name: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell (default A1)
        end_cell: Ending cell (optional, auto-expands if not provided)
        preview_only: Whether to return preview only
    
    Returns:  
    JSON string containing structured cell data with validation metadata.
    Each cell includes: address, value, row, column, and validation info (if any).
    """
    try:
        full_path = get_excel_path(file_name, user_id)
        from utils.data import read_excel_range_with_metadata
        result = read_excel_range_with_metadata(
            full_path, 
            sheet_name, 
            start_cell, 
            end_cell
        )
        if not result or not result.get("cells"):
            return "No data found in specified range"
            
        # Return as formatted JSON string
        import json
        return json.dumps(result, indent=2, default=str)
        
    except Exception as e:
        logger.error(f"Error reading data: {e}")
        raise

@mcp.tool()
def write_data_to_excel(
    user_id: str,
    file_name: str,
    sheet_name: str,
    data: List[List],
    start_cell: str = "A1",
) -> str:
    """
    Write data to Excel worksheet.
    Excel formula will write to cell without any verification.

    PARAMETERS:  
    user_id: User ID for file organization
    file_name: Path to Excel file
    sheet_name: Name of worksheet to write to
    data: List of lists containing data to write to the worksheet, sublists are assumed to be rows
    start_cell: Cell to start writing to, default is "A1"
  
    """
    try:
        full_path = get_excel_path(file_name, user_id)
        result = write_data(full_path, sheet_name, data, start_cell)
        return result["message"]
    except (ValidationError, DataError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error writing data: {e}")
        raise

@mcp.tool()
def create_workbook(user_id: str, file_name: str) -> str:
    """Create new Excel workbook."""
    try:
        full_path = get_excel_path(file_name, user_id)
        from utils.workbook import create_workbook as create_workbook_impl
        create_workbook_impl(full_path)
        return f"Created workbook at {full_path}"
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating workbook: {e}")
        raise

@mcp.tool()
def create_worksheet(user_id: str, file_name: str, sheet_name: str) -> str:
    """Create new worksheet in workbook."""
    try:
        full_path = get_excel_path(file_name, user_id)
        from utils.workbook import create_sheet as create_worksheet_impl
        result = create_worksheet_impl(full_path, sheet_name)
        return result["message"]
    except (ValidationError, WorkbookError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating worksheet: {e}")
        raise

@mcp.tool()
def create_chart(
    user_id: str,
    file_name: str,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis: str = "",
    y_axis: str = ""
) -> str:
    """Create chart in worksheet."""
    try:
        full_path = get_excel_path(file_name, user_id)
        result = create_chart_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            chart_type=chart_type,
            target_cell=target_cell,
            title=title,
            x_axis=x_axis,
            y_axis=y_axis
        )
        return result["message"]
    except (ValidationError, ChartError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating chart: {e}")
        raise

@mcp.tool()
def create_pivot_table(
    user_id: str,
    file_name: str,
    sheet_name: str,
    data_range: str,
    rows: List[str],
    values: List[str],
    columns: Optional[List[str]] = None,
    agg_func: str = "sum"
) -> str:
    """Create pivot table in worksheet."""
    try:
        full_path = get_excel_path(file_name, user_id)
        result = create_pivot_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            rows=rows,
            values=values,
            columns=columns or [],
            agg_func=agg_func
        )
        return result["message"]
    except (ValidationError, PivotError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating pivot table: {e}")
        raise

@mcp.tool()
def create_table(
    user_id: str,
    file_name: str,
    sheet_name: str,
    data_range: str,
    table_name: Optional[str] = None,
    table_style: str = "TableStyleMedium9"
) -> str:
    """Creates a native Excel table from a specified range of data."""
    try:
        full_path = get_excel_path(file_name, user_id)
        result = create_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            table_name=table_name,
            table_style=table_style
        )
        return result["message"]
    except DataError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating table: {e}")
        raise

@mcp.tool()
def copy_worksheet(
    user_id: str,
    file_name: str,
    source_sheet: str,
    target_sheet: str
) -> str:
    """Copy worksheet within workbook."""
    try:
        full_path = get_excel_path(file_name, user_id)
        result = copy_sheet(full_path, source_sheet, target_sheet)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying worksheet: {e}")
        raise

@mcp.tool()
def delete_worksheet(
    user_id: str,
    file_name: str,
    sheet_name: str
) -> str:
    """Delete worksheet from workbook."""
    try:
        full_path = get_excel_path(file_name, user_id)
        result = delete_sheet(full_path, sheet_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting worksheet: {e}")
        raise

@mcp.tool()
def rename_worksheet(
    user_id: str,
    file_name: str,
    old_name: str,
    new_name: str
) -> str:
    """Rename worksheet in workbook."""
    try:
        full_path = get_excel_path(file_name, user_id)
        result = rename_sheet(full_path, old_name, new_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error renaming worksheet: {e}")
        raise

@mcp.tool()
def get_workbook_metadata(
    user_id: str,
    file_name: str,
    include_ranges: bool = False
) -> str:
    """Get metadata about workbook including sheets, ranges, etc."""
    try:
        full_path = get_excel_path(file_name, user_id)
        result = get_workbook_info(full_path, include_ranges=include_ranges)
        return str(result)
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting workbook metadata: {e}")
        raise

@mcp.tool()
def merge_cells(user_id: str, file_name: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """Merge a range of cells."""
    try:
        full_path = get_excel_path(file_name, user_id)
        result = merge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error merging cells: {e}")
        raise

@mcp.tool()
def unmerge_cells(user_id: str, file_name: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """Unmerge a range of cells."""
    try:
        full_path = get_excel_path(file_name, user_id)
        result = unmerge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error unmerging cells: {e}")
        raise

@mcp.tool()
def get_merged_cells(user_id: str, file_name: str, sheet_name: str) -> str:
    """Get merged cells in a worksheet."""
    try:
        full_path = get_excel_path(file_name, user_id)
        return str(get_merged_ranges(full_path, sheet_name))
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting merged cells: {e}")
        raise

@mcp.tool()
def copy_range(
    user_id: str,
    file_name: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    target_sheet: Optional[str] = None
) -> str:
    """Copy a range of cells to another location."""
    try:
        full_path = get_excel_path(file_name, user_id)
        from utils.sheet import copy_range_operation
        result = copy_range_operation(
            full_path,
            sheet_name,
            source_start,
            source_end,
            target_start,
            target_sheet or sheet_name  # Use source sheet if target_sheet is None
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying range: {e}")
        raise

@mcp.tool()
def delete_range(
    user_id: str,
    file_name: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    shift_direction: str = "up"
) -> str:
    """Delete a range of cells and shift remaining cells."""
    try:
        full_path = get_excel_path(file_name, user_id)
        from utils.sheet import delete_range_operation
        result = delete_range_operation(
            full_path,
            sheet_name,
            start_cell,
            end_cell,
            shift_direction
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting range: {e}")
        raise

@mcp.tool()
def validate_excel_range(
    user_id: str,
    file_name: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None
) -> str:
    """Validate if a range exists and is properly formatted."""
    try:
        full_path = get_excel_path(file_name, user_id)
        range_str = start_cell if not end_cell else f"{start_cell}:{end_cell}"
        result = validate_range_impl(full_path, sheet_name, range_str)
        return result["message"]
    except ValidationError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating range: {e}")
        raise

@mcp.tool()
def get_data_validation_info(
    user_id: str,
    file_name: str,
    sheet_name: str
) -> str:
    """
    Get all data validation rules in a worksheet.
    
    This tool helps identify which cell ranges have validation rules
    and what types of validation are applied.
    
    Args:
        user_id: User ID for file organization
        file_name: Path to Excel file
        sheet_name: Name of worksheet
        
    Returns:
        JSON string containing all validation rules in the worksheet
    """
    try:
        full_path = get_excel_path(file_name, user_id)
        from openpyxl import load_workbook
        from utils.cell_validation import get_all_validation_ranges
        
        wb = load_workbook(full_path, read_only=False)
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found"
            
        ws = wb[sheet_name]
        validations = get_all_validation_ranges(ws)
        wb.close()
        
        if not validations:
            return "No data validation rules found in this worksheet"
            
        import json
        return json.dumps({
            "sheet_name": sheet_name,
            "validation_rules": validations
        }, indent=2, default=str)
        
    except Exception as e:
        logger.error(f"Error getting validation info: {e}")
        raise
    
@mcp.tool()
def list_minio_files(
    user_id: str
    ):
    # this function will list all the files in the minio bucket, under the /bucket_name/private/user_id/ folder
    # and return a list of file names
    configs = load_configs()
    endpoint = configs.MINIO_ENDPOINT
    access_key = configs.MINIO_ACCESS_KEY
    secret_key = configs.MINIO_SECRET_KEY
    bucket_name = configs.MINIO_BUCKET  # should be 'ai-file'
    prefix = f"private/{user_id}/"      # per-user folder in 'private/'
    
    client = Minio(
        endpoint.replace("http://", "").replace("https://", ""),
        access_key=access_key,
        secret_key=secret_key,
        secure=False  # Set to False since endpoint uses http://
    )
    
    try:
        objects = client.list_objects(bucket_name, prefix=prefix, recursive=True)
        file_list = []
        for obj in objects:
            file_info = {
                "name": obj.object_name,
                "size": obj.size,
                "last_modified": obj.last_modified.isoformat() if obj.last_modified else None
            }
            file_list.append(file_info)
            logger.info(f"Found file: {obj.object_name}")
        
        return file_list
    except Exception as e:
        logger.error(f"Error listing MinIO files: {e}")
        raise

@mcp.tool()
def pull_minio_file(
    user_id: str,
    file_name: str,
):
    """
    Download a file from MinIO to the local Excel directory.

    Source  :     <bucket>/private/{user_id}/{file_name}
    Target  :     <EXCEL_FILES_PATH>/{user_id}/{file_name}

    Args:
        user_id:   ID of the user (used for the per-user folder)
        file_name: Name of the file in MinIO
    Returns:
        Absolute path (str) of the downloaded local file
    """
    configs = load_configs()
    client = Minio(
        configs.MINIO_ENDPOINT.replace("http://", "").replace("https://", ""),
        access_key=configs.MINIO_ACCESS_KEY,
        secret_key=configs.MINIO_SECRET_KEY,
        secure=False,  # Endpoint is plain HTTP in most dev setups
    )
    bucket_name = configs.MINIO_BUCKET

    # Define the object path in MinIO and local file path
    object_name = f"private/{user_id}/{file_name}"
    local_file_path = Path(get_excel_path(file_name, user_id))

    # Create local directory if it doesn't exist
    local_file_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        # Download the file from MinIO
        client.fget_object(bucket_name, object_name, str(local_file_path))
        logger.info(f"Successfully pulled file {object_name} from MinIO to {local_file_path}")
        return str(local_file_path)
    except S3Error as e:
        logger.error(f"Error pulling file from MinIO: {e}")
        raise

@mcp.tool()
def push_minio_file(
    user_id: str,
    file_name: str,
):
    """
    Upload a local Excel file to MinIO, then remove the local copy.

    Source (local): <EXCEL_FILES_PATH>/{user_id}/{file_name}
    Dest   (MinIO): <bucket>/private/{user_id}/{file_name}(mod)

    The “(mod)” suffix differentiates uploaded files from their originals.
    """
    configs = load_configs()
    client = Minio(
        configs.MINIO_ENDPOINT.replace("http://", "").replace("https://", ""),
        access_key=configs.MINIO_ACCESS_KEY,
        secret_key=configs.MINIO_SECRET_KEY,
        secure=False,
    )
    bucket_name = configs.MINIO_BUCKET

    # Define local file path
    local_file_path = Path(get_excel_path(file_name, user_id))

    if not local_file_path.is_file():
        raise FileNotFoundError(f"File not found: {local_file_path}")

    # Create modified file name with (mod) extension
    modified_file_name = f"{local_file_path.stem}(mod){local_file_path.suffix}"
    object_name = f"private/{user_id}/{modified_file_name}"

    try:
        # Upload the file to MinIO
        client.fput_object(bucket_name, object_name, str(local_file_path))
        logger.info(f"Successfully pushed file {local_file_path} to MinIO as {object_name}")

        # Remove the local file after successful upload
        local_file_path.unlink()
        logger.info(f"Removed local file: {local_file_path}")

        return f"File uploaded to MinIO as {object_name}"
    except S3Error as e:
        logger.error(f"Error pushing file to MinIO: {e}")
        raise


def main():
    configs = load_configs()
    excel_files_path = configs.EXCEL_FILES_PATH
    port = configs.PORT
    host = configs.HOST
    log_level = configs.LOG_LEVEL
    
    # Create directory if it doesn't exist
    os.makedirs(excel_files_path, exist_ok=True)
    
    try:
        logger.info(f"Starting Excel MCP server (files directory: {excel_files_path})")
        logger.info(f"Host: {host}")
        logger.info(f"Port: {port}")
        logger.info(f"Log level: {log_level}")
        mcp.run(
            transport="http",
            host=host,
            port=port,
            log_level=log_level,
        )
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")
        
if __name__ == "__main__":
    main()