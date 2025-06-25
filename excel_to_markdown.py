#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel转Markdown工具
支持多工作表、中文内容、自定义选项
"""

import pandas as pd
import openpyxl
import argparse
import os
import sys
from pathlib import Path
from typing import List, Optional, Dict, Any


class ExcelToMarkdownConverter:
    """Excel转Markdown转换器"""

    def __init__(self, file_path: str):
        """
        初始化转换器

        Args:
            file_path: Excel文件路径
        """
        self.file_path = Path(file_path)
        self.workbook = None
        self.sheet_names = []

        if not self.file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        if not self.file_path.suffix.lower() in ['.xlsx', '.xls']:
            raise ValueError("请提供Excel文件(.xlsx或.xls)")

        self._load_workbook()

    def _load_workbook(self):
        """加载Excel工作簿"""
        try:
            # 使用openpyxl读取工作表名称
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            self.sheet_names = wb.sheetnames
            wb.close()
            print(f"成功加载文件: {self.file_path}")
            print(f"发现工作表: {', '.join(self.sheet_names)}")
        except Exception as e:
            raise Exception(f"加载Excel文件失败: {str(e)}")

    def read_sheet(self, sheet_name: Optional[str] = None, **kwargs) -> pd.DataFrame:
        """
        读取指定工作表

        Args:
            sheet_name: 工作表名称，如果为None则读取第一个工作表
            **kwargs: pandas.read_excel的其他参数

        Returns:
            DataFrame对象
        """
        if sheet_name is None:
            sheet_name = self.sheet_names[0]

        if sheet_name not in self.sheet_names:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")

        try:
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                header=None,  # 不自动设置表头
                **kwargs
            )
            return df
        except Exception as e:
            raise Exception(f"读取工作表 '{sheet_name}' 失败: {str(e)}")

    def dataframe_to_markdown(self,
                              df: pd.DataFrame,
                              include_headers: bool = True,
                              skip_empty_rows: bool = True,
                              header_row: int = 0) -> str:
        """
        将DataFrame转换为Markdown表格

        Args:
            df: DataFrame对象
            include_headers: 是否包含表头
            skip_empty_rows: 是否跳过空行
            header_row: 表头行索引（当include_headers=True时）

        Returns:
            Markdown格式的字符串
        """
        if df.empty:
            return "**该工作表为空**\n\n"

        # 处理空值
        df = df.fillna('')

        # 转换为字符串并清理
        df = df.astype(str).applymap(lambda x: x.strip() if x != 'nan' else '')

        # 跳过空行
        if skip_empty_rows:
            # 找出不全为空的行
            non_empty_mask = df.apply(lambda row: row.str.strip().ne('').any(), axis=1)
            df = df[non_empty_mask]

        if df.empty:
            return "**该工作表没有有效数据**\n\n"

        markdown_lines = []

        if include_headers and len(df) > header_row:
            # 使用指定行作为表头
            headers = df.iloc[header_row].tolist()
            headers = [str(h) if h else ' ' for h in headers]

            # 表头行
            markdown_lines.append('| ' + ' | '.join(headers) + ' |')
            # 分隔行
            markdown_lines.append('| ' + ' | '.join(['---'] * len(headers)) + ' |')

            # 数据行（跳过表头行）
            data_df = df.iloc[header_row + 1:]
        else:
            # 不包含表头，第一行数据后添加分隔符
            data_df = df
            if not data_df.empty:
                first_row = data_df.iloc[0].tolist()
                first_row = [str(cell) if cell else ' ' for cell in first_row]
                markdown_lines.append('| ' + ' | '.join(first_row) + ' |')
                markdown_lines.append('| ' + ' | '.join(['---'] * len(first_row)) + ' |')
                data_df = data_df.iloc[1:]

        # 处理数据行
        for _, row in data_df.iterrows():
            cells = [str(cell) if cell else ' ' for cell in row.tolist()]
            markdown_lines.append('| ' + ' | '.join(cells) + ' |')

        return '\n'.join(markdown_lines) + '\n\n'

    def convert_sheet(self,
                      sheet_name: Optional[str] = None,
                      include_headers: bool = True,
                      skip_empty_rows: bool = True,
                      header_row: int = 0) -> str:
        """
        转换单个工作表为Markdown

        Args:
            sheet_name: 工作表名称
            include_headers: 是否包含表头
            skip_empty_rows: 是否跳过空行
            header_row: 表头行索引

        Returns:
            Markdown格式的字符串
        """
        if sheet_name is None:
            sheet_name = self.sheet_names[0]

        print(f"正在转换工作表: {sheet_name}")

        # 读取数据
        df = self.read_sheet(sheet_name)

        # 生成Markdown
        markdown = f"# {sheet_name}\n\n"
        markdown += self.dataframe_to_markdown(
            df,
            include_headers=include_headers,
            skip_empty_rows=skip_empty_rows,
            header_row=header_row
        )

        return markdown

    def convert_all_sheets(self,
                           include_headers: bool = True,
                           skip_empty_rows: bool = True,
                           header_row: int = 0) -> str:
        """
        转换所有工作表为Markdown

        Args:
            include_headers: 是否包含表头
            skip_empty_rows: 是否跳过空行
            header_row: 表头行索引

        Returns:
            包含所有工作表的Markdown字符串
        """
        all_markdown = []

        for sheet_name in self.sheet_names:
            sheet_markdown = self.convert_sheet(
                sheet_name=sheet_name,
                include_headers=include_headers,
                skip_empty_rows=skip_empty_rows,
                header_row=header_row
            )
            all_markdown.append(sheet_markdown)

        return '\n---\n\n'.join(all_markdown)

    def save_markdown(self,
                      markdown_content: str,
                      output_path: Optional[str] = None) -> str:
        """
        保存Markdown内容到文件

        Args:
            markdown_content: Markdown内容
            output_path: 输出文件路径，如果为None则自动生成

        Returns:
            输出文件路径
        """
        if output_path is None:
            output_path = self.file_path.with_suffix('.md')

        output_path = Path(output_path)

        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(markdown_content)
            print(f"Markdown文件已保存: {output_path}")
            return str(output_path)
        except Exception as e:
            raise Exception(f"保存文件失败: {str(e)}")


def main():
    """命令行接口"""
    parser = argparse.ArgumentParser(description='Excel转Markdown工具')
    parser.add_argument('input_file', help='输入的Excel文件路径')
    parser.add_argument('-o', '--output', help='输出的Markdown文件路径')
    parser.add_argument('-s', '--sheet', help='指定工作表名称（默认转换所有工作表）')
    parser.add_argument('--no-headers', action='store_true', help='不包含表头')
    parser.add_argument('--keep-empty', action='store_true', help='保留空行')
    parser.add_argument('--header-row', type=int, default=0, help='表头行索引（默认0）')
    parser.add_argument('--list-sheets', action='store_true', help='列出所有工作表名称')

    args = parser.parse_args()

    try:
        # 创建转换器
        converter = ExcelToMarkdownConverter(args.input_file)

        # 如果只是列出工作表
        if args.list_sheets:
            print("工作表列表:")
            for i, sheet_name in enumerate(converter.sheet_names, 1):
                print(f"  {i}. {sheet_name}")
            return

        # 转换参数
        include_headers = not args.no_headers
        skip_empty_rows = not args.keep_empty

        # 执行转换
        if args.sheet:
            # 转换指定工作表
            markdown_content = converter.convert_sheet(
                sheet_name=args.sheet,
                include_headers=include_headers,
                skip_empty_rows=skip_empty_rows,
                header_row=args.header_row
            )
        else:
            # 转换所有工作表
            markdown_content = converter.convert_all_sheets(
                include_headers=include_headers,
                skip_empty_rows=skip_empty_rows,
                header_row=args.header_row
            )

        # 保存结果
        output_path = converter.save_markdown(markdown_content, args.output)
        print(f"转换完成! 输出文件: {output_path}")

    except Exception as e:
        print(f"错误: {str(e)}", file=sys.stderr)
        sys.exit(1)


# 使用示例
if __name__ == '__main__':
    # 如果直接运行脚本
    if len(sys.argv) == 1:
        # 交互式模式示例
        print("Excel转Markdown工具 - 交互式模式")
        print("=" * 50)

        # 示例用法
        example_file = "MCIO 74PIN VT（三合一，CAP不贴Mylar）SOPA8， 中越版.xlsx"

        try:
            if os.path.exists(example_file):
                converter = ExcelToMarkdownConverter(example_file)

                print(f"发现文件: {example_file}")
                print("工作表列表:")
                for i, sheet_name in enumerate(converter.sheet_names, 1):
                    print(f"  {i}. {sheet_name}")

                # 转换第一个工作表
                markdown_content = converter.convert_sheet(
                    include_headers=True,
                    skip_empty_rows=True
                )

                # 保存文件
                output_path = converter.save_markdown(markdown_content)
                print(f"\n转换完成! 查看文件: {output_path}")

            else:
                print("使用方法:")
                print("1. 命令行模式: python excel_to_markdown.py <excel_file>")
                print("2. 编程模式:")
                print("   converter = ExcelToMarkdownConverter('file.xlsx')")
                print("   markdown = converter.convert_all_sheets()")
                print("   converter.save_markdown(markdown)")

        except Exception as e:
            print(f"示例运行失败: {str(e)}")
            print("\n请确保安装了必要的依赖:")
            print("pip install pandas openpyxl")
    else:
        # 命令行模式
        main()